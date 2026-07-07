import os
import re
import openpyxl

desktop_path = r"C:\Users\JesusAlexisCarmonaCa\Desktop"
email_regex = re.compile(r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}')

print("Searching Desktop files for email addresses...")

# 1. Search text files, csv, env
files_to_check = [
    "email.csv",
    "emailimporttertools.py",
    "jesuscarm.o.n.a@atomicmail.io.txt",
    "Ms enlaces a refere manus i.txt"
]

for filename in os.listdir(desktop_path):
    filepath = os.path.join(desktop_path, filename)
    if os.path.isdir(filepath):
        continue
    
    # Check text/csv files
    if filename.endswith(('.txt', '.csv', '.py', '.env', '.local', '.json', '.md')):
        try:
            with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
                content = f.read()
                emails = email_regex.findall(content)
                if emails:
                    unique_emails = list(set(emails))
                    print(f"File: {filename} -> found: {unique_emails[:5]}")
        except Exception as e:
            pass

# 2. Check Excel files
excel_files = [
    r"cuentas lovable platzi supabase\lovable ventas\credenciales supabase y lovable.xlsx",
    r"cuentas lovable platzi supabase\lovable ventas\emaillovable.xlsx"
]

for rel_path in excel_files:
    filepath = os.path.join(desktop_path, rel_path)
    if os.path.exists(filepath):
        try:
            wb = openpyxl.load_workbook(filepath, read_only=True)
            emails_found = []
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                for row in ws.iter_rows(values_only=True):
                    for cell in row:
                        if cell and isinstance(cell, str):
                            matches = email_regex.findall(cell)
                            if matches:
                                emails_found.extend(matches)
            if emails_found:
                print(f"Excel: {os.path.basename(filepath)} -> found: {list(set(emails_found))[:10]}")
        except Exception as e:
            print(f"Error reading Excel {filename}: {e}")
